import json

import openpyxl as pyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import re


class ExcelIO:
    worksheet: Worksheet
    workbook: Workbook
    alphabet = list(chr(n) for n in range(65, 91))

    def __init__(self, file_name=None):
        try:
            self.workbook = pyxl.load_workbook(f'{file_name}.xlsx')
        finally:
            pass
        self.name = str
        self.worksheet = ...

    def set_workbook(self, file_path):
        self.workbook = pyxl.load_workbook(file_path)
        self.name = file_path
        self.worksheet = self.workbook[self.workbook.sheetnames[0]]

    @staticmethod
    def get_cell(x: int, y: int, letter: str = '') -> str:

        while x >= 1:
            if x <= 26:
                letter += ExcelIO.alphabet[x - 1]
                x = 0
            else:
                letter += ExcelIO.alphabet[int(x) // 26 - 1]
                x %= 26
        return letter + str(y)

    def extract_headers(self, sheet_number):
        self.worksheet = self.workbook[self.workbook.sheetnames[sheet_number]]
        export_col = 1
        excel_data = []
        val = self.worksheet[self.get_cell(export_col, 1)].value
        while val is not None:
            excel_data.append(val)
            export_col += 1
            val = self.worksheet[self.get_cell(export_col, 1)].value
        return excel_data

    def extract_data(self, sheet_number):
        self.worksheet = self.workbook[self.workbook.sheetnames[sheet_number]]
        excel_data = []
        export_row = 1
        export_col = 1
        while self.worksheet[self.get_cell(1, export_row)].value is not None:
            export_row += 1
        while self.worksheet[self.get_cell(export_col, 1)].value is not None:
            export_col += 1
        print(export_col, export_row)
        for r in range(2, export_row):
            excel_data.append([])
            for c in range(1, export_col):
                val = self.worksheet[self.get_cell(c, r)].value
                try:
                    val = int(val)
                except ValueError:
                    pass
                except TypeError:
                    pass
                finally:
                    excel_data[r - 2].append(val)
        return excel_data


class SqlIO:
    def __init__(self, file_name):
        self.file_name = file_name
        self.file = open(f'{file_name}.sql', 'w')

    def json_to_table_values(self, name, rows):
        self.file.write(f'INSERT INTO {name} VALUES\n')
        for row in rows:
            self.file.write(str(tuple(row)).
                            replace("None", 'NULL').
                            replace("'NULL'", 'NULL').
                            replace(",)", ")"))
            self.file.write(",\n" if row != rows[-1] else ";\n\n")

    def close(self):
        self.file.close()


class JsonIO:
    data: list[dict]

    def __init__(self, file_path):
        self.file_path = file_path
        self.file = open(self.file_path, 'w', encoding="utf-8", errors="replace")
        self.data = None
        self.file.close()

    def add_entries(self, txt, keys):

        with open(self.file_path, 'r', encoding="utf-8", errors="replace") as self.file:
            self.data = []
        print(txt)
        for n, txt_input in enumerate(txt):
            print(txt_input)
            entry = {}
            for index, key in enumerate(keys):
                print(key, txt_input[index])
                entry[key] = txt_input[index]
            self.data.append(entry)

        self.file = open(self.file_path, 'w', encoding="utf-8", errors="replace")
        self.file.write(json.dumps(self.data, sort_keys=False, indent=4))
        self.file.close()

    def add_field(self, intended, external, *combination):
        with open(self.file_path, 'r') as self.file:
            self.data = json.load(self.file)
        for n in range(len(self.data)):
            placeholder = []
            for combo in combination:
                if self.data[n][combo] is None:
                    continue
                else:
                    placeholder.append(self.data[n][combo])
            print("PLACEHOLDER:", placeholder)
            self.data[n][intended] = self.find_identical(intended, placeholder, external)
        self.file = open(self.file_path, 'w', encoding="utf-8", errors="replace")
        self.file.write(json.dumps(self.data, sort_keys=False, indent=4))
        self.file.close()

    @staticmethod
    def find_identical(field, this_data, other_data):
        for row in other_data:
            found = True
            for item in this_data:
                # print(item)
                if item not in row.values():
                    found = False
            if found:
                print("FOUND?")
                return row[field]

    def clear_entries(self):
        with open(self.file_path, 'w', encoding="utf-8", errors="replace") as self.file:
            self.file.write(json.dumps([], sort_keys=False, indent=4))

    def get_entries(self):
        try:
            with open(self.file_path, 'r') as self.file:
                self.data = json.load(self.file)
                print("DATA:", self.data)
        except FileNotFoundError:
            print("NOT FOUND?")
        return self.data

    def get_keys(self):
        with open(self.file_path, 'r') as self.file:
            self.data = json.load(self.file)
        print("KEYS")
        return tuple(self.data[0])

    def get_columns(self, *args, **kwargs):
        cols = []
        with open(self.file_path, 'r') as self.file:
            self.data = json.load(self.file)
        for dictionary in self.data:
            cont = True
            for flag, req in kwargs.items():
                print("REQ:", dictionary[req if type(req) == str else req[0]], "| OK:", NONE_TYPES)
                if 'req' in flag:
                    if dictionary[req if type(req) == str else req[0]] in NONE_TYPES:
                        cont = False
                        break
                elif 'null' in flag:
                    if dictionary[req] not in NONE_TYPES:
                        cont = False
                        break
                elif flag == 'unique':
                    for row in cols:
                        if dictionary[req] in row:
                            cont = False
                            break
                elif 'equals' in flag:
                    if dictionary[req[0]] != req[1]:
                        cont = False
                    else:
                        print('Equals checks out.', dictionary[req[0]], req[1])
                elif 'multi' in flag:
                    norm_list = []
                    mult_list = []
                    for name in args:
                        if name == req:
                            mult_list = dictionary[req].replace(" ", "").split(',')
                            print("MULTI:", mult_list)

                        else:
                            try:
                                norm_list.append(int(dictionary[name]))
                            except TypeError:
                                norm_list.append(dictionary[name])
                            except ValueError:
                                norm_list.append(dictionary[name])
                            except KeyError:
                                pass
                    for loc, item in enumerate(mult_list):
                        try:
                            item = int(item)
                        except ValueError or TypeError:
                            pass
                        if 'special' in flag:
                            temp = []
                            item = re.split(':|-->|mi', item)
                            temp2 = []
                            multi_count = len(args) - args.index(req) - (1 if 'index' in flag else 0)
                            for n in range(multi_count):
                                temp2.append(item[n])
                            item = temp2
                            print("SPLIT:", item)
                            for sub in item:
                                try:
                                    sub = int(sub)
                                except ValueError or TypeError:
                                    pass
                                temp.append(sub)
                            item = temp
                        else:
                            item = [item]
                        is_unique = True
                        if 'special' in flag:
                            for col in cols:
                                if col[0] == item[0]:
                                    is_unique = False
                        if not is_unique:
                            continue
                        if 'index' in flag:
                            cols.append(norm_list + [*item, loc])
                        else:
                            cols.append(norm_list + [*item])
                    cont = False

            if not cont:
                print("SKIPPING")
                continue
            cols.append([])
            for name in args:
                cols[-1].append(dictionary[name])
        return cols


json_data = {0: JsonIO("output/json/airplanes.json"),
             1: JsonIO("output/json/airports.json"),
             2: JsonIO("output/json/locations.json"),
             3: JsonIO("output/json/persons.json"),
             4: JsonIO("output/json/flights.json"),
             5: JsonIO("output/json/routes.json")}

NONE_TYPES = (None, 'null', 'NULL', 'None')

tables = (('person', 3, ('personID', 'first_name', 'last_name', 'locationID'), {}),
          ('pilot', 3, ('personID', 'taxID', 'experience', 'flightID'), dict(req1="taxID")),
          ('pilot_license', 3, ('personID', 'license_types'), dict(req1="taxID", multi='license_types')),
          ('passenger', 3, ('personID', 'miles', 'funds'), dict(null='taxID')),
          ('passenger_vacation', 3, ('personID', 'vacations', 'sequence'), dict(req='vacations', null='taxID', multi_index='vacations')),
          ('airline', 0, ('airlineID', 'airline_revenue'), dict(unique='airlineID')),
          ('flight', 4, ('flightID', 'cost', 'routeID', 'support_airline', 'support_tail', 'progress', 'airplane_status', 'next_time'), {}),
          ('route', 5, ('routeID',), {}),
          ('leg', 5, ('legs', 'distance', 'arrives_at', 'departs_from'), dict(multi_special='legs')),
          ('airport', 1, ('airportID', 'airport_name', 'city', 'state', 'country_code', 'locationID'), {}),
          ('location', 2, ('locationID',), {}),
          ('airplane', 0, ('airlineID', 'tail_num', 'seat_capacity', 'speed', 'locationID'), {}),
          ('prop', 0, ('airlineID', 'tail_num', 'skids', 'props'), dict(equals=('plane_type', 'prop'))),
          ('jet', 0, ('airlineID', 'tail_num', 'jets'), dict(equals=('plane_type', 'jet'))),
          ('contains', 5, ('routeID', 'legs', 'sequence'), dict(multi_special_index='legs')))

if __name__ == '__main__':
    spreadsheet = ExcelIO('data')
    for n in range(6):
        json_data[n].add_entries(spreadsheet.extract_data(n), spreadsheet.extract_headers(n))
    json_data[3].add_field('flightID', json_data[4].get_entries(), 'flying_airline', 'flying_tail')
    database = SqlIO('output/output')
    for title, table, columns, reqs in tables:
        database.json_to_table_values(title, json_data[table].get_columns(*columns, **reqs))
    database.close()
